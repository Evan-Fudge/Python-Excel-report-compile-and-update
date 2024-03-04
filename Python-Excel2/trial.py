import openpyxl
from openpyxl.styles import Font

master_data= openpyxl.load_workbook('master_data_sheet.xlsx')
daily_data= openpyxl.load_workbook('daily_sheet.xlsx')

master_sheet = master_data['data']
daily_sheet = daily_data['Sheet1']



# This is the formula that gets us our daily_sheet's row count with data in the cells.
is_data = True
daily_row_count = 1

while is_data:
    daily_row_count += 1
    data = daily_sheet.cell(row= daily_row_count, column=1).value
    if data == None:
        is_data = False
#print(daily_row_count)

# This gets us our row count for master sheet with data in cells

is_data = True
master_row_count = 1

while is_data:
    master_row_count += 1
    data = master_sheet.cell(row= master_row_count, column=1).value
    if data == None:
        is_data = False

#print(master_row_count)

# we now must extract data from daily sheet
# we need to extract this data and store it into a list of dictionaries

todays_data = []
for i in range ( 1, daily_row_count):
    row_data = {}
    row_data['ID'] = daily_sheet.cell(row= i, column= 1 ).value
    row_data['TODAY_PURCHASES'] = daily_sheet.cell(row=i, column=2).value
    row_data['TODAYS_REWARDS'] = daily_sheet.cell(row=i, column=3).value
    todays_data.append(row_data)
#print(todays_data)
#[{'ID': 'ID', 'TODAY_PURCHASES': 'TODAY PURCHASES', 'TODAYS_REWARDS': 'TODAYS REWARDS'}

# Write daily sheet into master sheet
# Find row using id column
# Go to total purchases cell and add in today's purchases
# Go to total reward balance and add today's reward

for i in range(2,master_row_count):
    id = master_sheet.cell(row= i, column=1).value
    for row in todays_data:
        if row['ID'] == id:
            TODAYS_PURCHASES = row['TODAY_PURCHASES']
            TODAYS_REWARDS = row['TODAYS_REWARDS']

            # Get data from master sheet
            TOTAL_PURCHASES = master_sheet.cell(row=i, column=5).value
            TOTAL_REWARDS = master_sheet.cell(row=i, column=6).value

            # Add values of today's data into total data

            new_total_purchase = TODAYS_PURCHASES + TOTAL_PURCHASES
            new_total_rewards =  TODAYS_REWARDS + TOTAL_REWARDS

            master_sheet.cell(row=i, column=5).value = new_total_purchase
            master_sheet.cell(row=i, column=6).value = new_total_rewards

for letter in ['A','B','C','D','E','F']:
   max_width = 0

   for row_number in range(1, master_sheet.max_row +1):
      if len(str(master_sheet[f'{letter}{row_number}'].value)) > max_width:
         max_width = len(str(master_sheet[f'{letter}{row_number}'].value))

   master_sheet.column_dimensions[letter].width = max_width + 10

master_data.save('New_Master_Report.xlsx')

# Now we are going to create a daily report that contains the missing data that is included in the master data report but is missing from the daily report

#This creates the new report file that we are going to save to

daily_report = openpyxl.Workbook()
ws = daily_report.active

# Now we are going to get our headers to style them
is_data = True
column_count= 1
header_values = []

while is_data:
    column_count +=1
    data = master_sheet.cell(row=1, column= column_count).value
    if data != None:
        header_values.append(data)
    else:
        is_data = False


header_style = Font(name= 'Times New Roman', size= 12, bold=True)
for i, col_name in enumerate(header_values):
    col_index = i+1
    ws.cell(row=1, column=col_index).value = col_name
    ws.cell(row=1, column=col_index).font = header_style

IDs = []
for data in todays_data:
    IDs.append(data['ID'])
IDs.pop(0)
#print(IDs)

final_data = []
for i in range (2, master_row_count):
    ID = master_sheet.cell(row=i, column=1).value
    if ID in IDs:
        lst = []
        for j in range(2,7):
            lst.append(master_sheet.cell(row=i, column=j).value)
        final_data.append(lst)

for data in final_data:
    ws.append(data)

#daily_report.save('New_Daily_Report.xlsx')

for letter in ['A','B','C','D','E']:
   max_width = 0

   for row_number in range(1, ws.max_row +1):
      if len(str(ws[f'{letter}{row_number}'].value)) > max_width:
         max_width = len(str(ws[f'{letter}{row_number}'].value))

   ws.column_dimensions[letter].width = max_width + 10

daily_report.save('New_Daily_Report.xlsx')


#print(header_values)

